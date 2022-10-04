

#Excel utils for DataDriven testing usage:

import openpyxl                                     #openpyxl is used for read,write excel sheets
from openpyxl.styles import PatternFill             #solid color module package need to download if we need this


def getRowCount(file,sheetname):                        #To get row count from the excel sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getColumnCount(file,sheetname):                        #To get column count from the excel sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def readData(file,sheetname,rownum,colnum):                  #To get data from the excel sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.cell(rownum,colnum).value)

def writeData(file,sheetname,rownum,colnum,data):           #To write data into the excel sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)

def fillGreenColor(file,sheetname,rownum,colnum):           #To fill green color into the excel sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenFill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')    #60b212 - for getting green color highlight in excel
    sheet.cell(rownum,colnum).fill=greenFill
    workbook.save(file)

def fillRedColor(file,sheetname,rownum,colnum):           #To fill red color into the excel sheet
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redFill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')      #ff0000 - for getting red color highlight in excel
    sheet.cell(rownum,colnum).fill=redFill
    workbook.save(file)





